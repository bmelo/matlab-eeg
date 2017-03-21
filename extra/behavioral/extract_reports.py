'''
IDOR
by Bruno Melo <bruno.melo@idor.org>
'''

import glob, os, csv, re
from openpyxl import Workbook, load_workbook

# Main function
def main():
    baseDir = r'..\..\..\..\..\RAW_DATA\EEG\\'
    pattLogs = r'SUBJ*\*.log'
    logs = sorted(glob.glob(baseDir+pattLogs))
    for log in logs:
        print(log)
        report = ReportUnity(log)
        report.writeExcel( 'REPORT.xlsx' )
    #LogsPRJ1411(baseDir + 'SVM').clearBooks(".*ROI").toXls('SVM-only-ROIs.xlsx')


# Class to export data to an ExcelFile
class ReportUnity:
    # Constructor
    def __init__( self, fileIn ):
        self.subjid = ''
        self.file = fileIn
        self.extractReports()

    # Extracts each form of an log file
    def extractReports( self ):
        rBegin = r'== FORM REPORT BEGIN =='
        rEnd = r'== FORM REPORT END =='
        fileContent = open( self.file, encoding="utf8" ).read()
        #print(fileContent)
        rex = re.compile( rBegin+r'(.*?)'+rEnd, re.S|re.M )
        self.reports = rex.findall( fileContent )
        self.reportsDict = ReportUnity.reportToDict( self.reports )

    # Convert text to dict
    def reportToDict( reports ):
        dictOut = []
        for report in reports:
            dictR = {}
            rex = re.compile(r'([\w_]+):\t(.*?)(?=\n$|\n[\w_]+:)', re.S|re.M)
            for elems in rex.findall(report):
                dictR[elems[0]] = elems[1]
            dictOut.append( dictR )
        return dictOut

    # Write logs in the last line of each sheet
    def writeExcel( self, xlsFile ):
        wb = load_workbook(filename = xlsFile)
        self.writeSheet( wb, 'INTENSITY', self.intensity());
        self.writeSheet( wb, 'MANTRAS', self.mantras());
        self.writeSheet( wb, 'CANSAÇO', self.tiredness());
        self.writeSheet( wb, 'CONCENTRAÇÃO', self.concentration());
        self.writeSheet( wb, 'OBSERVAÇÕES', self.observacoes());
        wb.save(xlsFile)

    def writeSheet(self, wb, sheetName, values):
        print(sheetName)
        ws = wb[sheetName]
        nrow = ReportUnity.nextLine( ws )
        self.writeLine( ws, nrow, values )
        
    # Find next line that is empty
    def nextLine( ws ):
        for nrow in range(3,25):
            cell = ws.cell(row = nrow, column = 1)
            if not cell.value:
                return nrow

    # Write line with passed values
    def writeLine( self, ws, nrow, values ):
        subjid = self.reportsDict[0]['subjid']
        values = [subjid] + values

        nC = 1
        for value in values:
            cell = ws.cell(row = nrow, column = nC)
            cell.value = value
            nC+=1

    # Extract values of keys, following order
    # reports [0 .. 9]
    def extractValues( self, keys, reports = list(range(0,10)), mapValues={} ):
        values = [];
        for nR in reports:
            # Checking if report exists
            if nR >= len(self.reportsDict):
                break

            for key in keys:
                try:
                    value = self.reportsDict[nR][key]
                    # Translate values to mapped value, if needed
                    if value in mapValues.keys():
                        value = mapValues[value]
                    values.append(value)
                except:
                    break
        return values
    
    # List values of intensity
    def intensity(self):
        return self.extractValues( ['intensidade_emocao', 'intensidade_neutro'], mapValues={
            'Muito Fraca':1,
            'Fraca':2,
            'Moderada':3,
            'Intensa':4,
            'Muito Intensa':5
        } )

    # List values for mantras
    def mantras(self):
        return self.extractValues( ['mantras_emocao', 'mantras_neutro'], mapValues={
            'Inútil':1,
            'Pouco útil':2,
            'Moderada':3,
            'Útil':4,
            'Muito Útil':5
        } )

    # List values for tiredness
    def tiredness(self):
        return self.extractValues( ['cansaco'], mapValues={
            'Nada':1,
            'Pouco':2,
            'Cansado':3,
            'Muito':4,
            'Extremamente':5
        } )


    # List values for tiredness
    def concentration(self):
        return self.extractValues( ['concentracao'], mapValues={
            'Nada':1,
            'Pouco':2,
            'Concentrado':3,
            'Muito':4,
            'Extremamente':5
        } )

    # List values of observations
    def observacoes(self):
        return self.extractValues( ['observacoes'] )

        
# Checks if this is the main file (when imported, is not used)
if __name__ == "__main__":
   main()